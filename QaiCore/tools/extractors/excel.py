"""
Extractor de datos de archivos Excel (.xlsx, .xls).

Dependencias: pip install openpyxl xlrd
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Any


def extract_data_from_excel(
    excel_path: str,
    sheet_name: str = None,
    as_dict: bool = True,
    header_row: int = None,
    skip_rows: int = 0
) -> Dict[str, Any] | List[List[Any]]:
    """
    Extrae datos de un archivo Excel.
    
    Args:
        excel_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja específica (None = hoja activa)
        as_dict: Si True, retorna {headers: [...], rows: [...]}, si False retorna lista de listas
        header_row: Índice de la fila que contiene los headers (0-indexed). 
                    Si es None, intenta detectarla automáticamente.
        skip_rows: Número de filas a saltar al inicio.
        
    Returns:
        Datos extraídos en formato dict o lista según as_dict
    """
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        
        # Extraer todas las filas como lista de listas
        all_rows = list(sheet.iter_rows(values_only=True))
        
        if not all_rows:
            return {"headers": [], "rows": []} if as_dict else []

        # 1. Aplicar skip_rows inicial
        rows_to_process = all_rows[skip_rows:]
        
        # 2. Detectar header_row si no se especifica
        # Buscamos palabras clave comunes en cartolas bancarias y tablas
        keywords = {'fecha', 'descripcion', 'descripción', 'monto', 'cargo', 'abono', 'saldo', 'total', 'rut', 'nro'}
        
        best_header_idx = 0
        max_matches = 0
        
        if header_row is None:
            # Buscamos en las primeras 50 filas
            for i, row in enumerate(rows_to_process[:50]):
                row_str = [str(c).lower() for c in row if c is not None]
                matches = [k for k in keywords if any(k in cell for cell in row_str)]
                num_matches = len(matches)
                
                # Bonus if it contains 'fecha' (very likely a table header)
                if any('fecha' in s for s in row_str):
                    num_matches += 2
                
                if num_matches > max_matches:
                    max_matches = num_matches
                    best_header_idx = i
            
            # Si no encontramos nada con al menos 2 matches, usar el fallback del inicio
            if max_matches < 2:
                for i, row in enumerate(rows_to_process[:10]):
                    if sum(1 for c in row if c is not None and str(c).strip() != "") > 2:
                        best_header_idx = i
                        break
            
            detected_header_idx = best_header_idx
        else:
            detected_header_idx = header_row

        # 3. Filtrar filas
        headers_raw = rows_to_process[detected_header_idx]
        data_rows = rows_to_process[detected_header_idx + 1:]
        
        # Limpiar filas vacías al final
        while data_rows and all(c is None or str(c).strip() == "" for c in data_rows[-1]):
            data_rows.pop()

        if as_dict:
            # Procesar headers: asegurar que sean strings y únicos
            headers = []
            for i, h in enumerate(headers_raw):
                name = str(h).strip() if h is not None else f"Col_{i}"
                if name == "" or name == "None":
                    name = f"Col_{i}"
                headers.append(name)
            
            result = {
                "sheet_name": sheet.title,
                "headers": headers,
                "rows": data_rows,
                "total_rows": len(data_rows),
                "header_row_index": detected_header_idx + skip_rows,
                "available_sheets": workbook.sheetnames
            }
            return result
        else:
            return rows_to_process[detected_header_idx:]
            
    except Exception as e:
        error_msg = f"❌ Error extrayendo datos del Excel: {str(e)}"
        return {"error": error_msg} if as_dict else [[error_msg]]


def excel_to_markdown_table(excel_path: str, sheet_name: str = None, max_rows: int = 50) -> str:
    """
    Convierte Excel a tabla Markdown para fácil lectura por agentes IA.
    """
    data = extract_data_from_excel(excel_path, sheet_name)
    
    if isinstance(data, dict) and "error" in data:
        return data["error"]
    
    headers = data["headers"]
    rows = data["rows"][:max_rows]
    
    # Construir tabla Markdown
    md_lines = [
        f"# Sheet: {data['sheet_name']}",
        f"*(Found headers at row {data['header_row_index']}. Showing {len(rows)} of {data['total_rows']} data rows)*\n",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |"
    ]
    
    for row in rows:
        row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
        md_lines.append(f"| {row_str} |")
    
    if data['total_rows'] > max_rows:
        md_lines.append(f"\n*... y {data['total_rows'] - max_rows} filas más*")
    
    return "\n".join(md_lines)


# Ejemplo de uso
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Uso: python excel.py <ruta_archivo.xlsx> [nombre_hoja]")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not Path(excel_file).exists():
        print(f"❌ Archivo no encontrado: {excel_file}")
        sys.exit(1)
    
    print(f"📊 Extrayendo datos de: {excel_file}")
    
    data = extract_data_from_excel(excel_file, sheet)
    
    if "error" in data:
        print(data["error"])
        sys.exit(1)

    print(f"\n✅ Hojas disponibles: {', '.join(data.get('available_sheets', []))}")
    print(f"✅ Hoja actual: {data.get('sheet_name', 'N/A')}")
    print(f"✅ Cabecera detectada en fila: {data.get('header_row_index', 0)}")
    print(f"✅ Total filas de datos: {data.get('total_rows', 0)}")
    
    print("\n" + "="*60)
    print("VISTA MARKDOWN (para agentes):")
    print("="*60)
    md_table = excel_to_markdown_table(excel_file, sheet, max_rows=15)
    print(md_table)

